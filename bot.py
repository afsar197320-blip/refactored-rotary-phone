import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
import threading
import os
import re
import time
import random
import requests
import ssl
import openpyxl
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED, as_completed
import urllib3

# ── Optional SSL workaround (kept in single file for VPS convenience) ─────
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
ssl._create_default_https_context = ssl._create_unverified_context

import telebot.apihelper as _apihelper
_tg_session = requests.Session()
_tg_session.verify = False
_apihelper.SESSION = _tg_session

# ============================================================
# CONFIG
# ============================================================
BOT_TOKEN = '8649575787:AAFgM8Eb5dVaumlTNTMzYxtu1eY3sN4YfkQ'
ADMIN_IDS = {8671204957}

DEFAULT_SERVER = 'limited.facebook.com'
DEFAULT_DEVICE = 'Random'
DEFAULT_BROWSER = 'Chrome'
DEFAULT_THREADS = 100
MAX_THREAD_LIMIT_NO_PROXY = 100
MAX_THREAD_LIMIT_WITH_PROXY = 300
PROGRESS_UPDATE_INTERVAL = 3
REQUEST_CONNECT_TIMEOUT = 8
REQUEST_READ_TIMEOUT = 15
MAX_RETRIES = 2
MIN_JITTER = 0.02
MAX_JITTER = 0.08
MAX_PROXY_TEST_WORKERS = 30
RESULTS_DIR = '.'

bot = telebot.TeleBot(BOT_TOKEN, parse_mode='HTML')


def is_admin(uid):
    if not ADMIN_IDS:
        return True
    return uid in ADMIN_IDS


# ============================================================
# SESSION STATE
# ============================================================
sessions = {}
session_lock = threading.Lock()


def create_default_session():
    return {
        'server': DEFAULT_SERVER,
        'device': DEFAULT_DEVICE,
        'browser': DEFAULT_BROWSER,
        'threads': DEFAULT_THREADS,
        'proxy_list': [],
        'numbers': [],
        'running': False,
        'stop_flag': False,
        'stats': {'checked': 0, 'found': 0, 'not_found': 0, 'error': 0},
        'progress_msg_id': None,
        'found_numbers': [],
        'not_found_numbers': [],
        'error_numbers': [],
        'awaiting': None,
        'last_progress_text': None,
        'last_progress_update': 0,
        'result_details': {},
        'start_ts': None,
        'active_workers': 0,
        'current_locale': 'en_US',
        'run_id': None,
        'last_result_paths': {},
    }


def get_session(cid):
    with session_lock:
        if cid not in sessions:
            sessions[cid] = create_default_session()
        return sessions[cid]


# ============================================================
# CORE DATA
# ============================================================
SERVER_MAP = {
    '1': 'm.facebook.com',
    '2': 'mbasic.facebook.com',
    '3': 'touch.facebook.com',
    '4': 'free.facebook.com',
    '5': 'm.alpha.facebook.com',
    '6': 'm.beta.facebook.com',
    '7': 'x.facebook.com',
    '8': DEFAULT_SERVER,
    '0': 'Random',
}

DEVICE_MAP = {
    '1': 'Android',
    '2': 'iPhone',
    '3': 'KaiOS',
    '4': 'Windows Phone',
    '5': 'BlackBerry',
    '0': 'Random',
}

BROWSER_MAP = {
    '1': 'Chrome', '2': 'Firefox', '3': 'Opera', '4': 'Edge', '5': 'Brave',
    '6': 'Samsung', '7': 'UC', '8': 'DuckDuckGo', '9': 'Vivaldi', '10': 'Yandex',
    '11': 'Kiwi', '12': 'Dolphin', '13': 'Mi Browser', '14': 'Maxthon', '15': 'Puffin',
    '0': 'Random',
}

COUNTRY_TO_LOCALE = {
    'AD': 'ca_ES', 'AE': 'ar_AR', 'AF': 'fa_IR', 'AG': 'en_US', 'AL': 'sq_AL',
    'AM': 'hy_AM', 'AO': 'pt_PT', 'AR': 'es_LA', 'AT': 'de_DE', 'AU': 'en_US',
    'AZ': 'az_AZ', 'BA': 'bs_BA', 'BB': 'en_US', 'BD': 'bn_IN', 'BE': 'nl_BE',
    'BF': 'fr_FR', 'BG': 'bg_BG', 'BH': 'ar_AR', 'BI': 'fr_FR', 'BJ': 'fr_FR',
    'BN': 'ms_MY', 'BO': 'es_LA', 'BR': 'pt_BR', 'BS': 'en_US', 'BT': 'en_US',
    'BW': 'en_US', 'BY': 'be_BY', 'BZ': 'en_US', 'CA': 'en_US', 'CD': 'fr_FR',
    'CF': 'fr_FR', 'CG': 'fr_FR', 'CH': 'de_DE', 'CI': 'fr_FR', 'CL': 'es_LA',
    'CM': 'fr_FR', 'CN': 'zh_CN', 'CO': 'es_LA', 'CR': 'es_LA', 'CU': 'es_LA',
    'CV': 'pt_PT', 'CY': 'el_GR', 'CZ': 'cs_CZ', 'DE': 'de_DE', 'DJ': 'fr_FR',
    'DK': 'da_DK', 'DM': 'en_US', 'DO': 'es_LA', 'DZ': 'ar_AR', 'EC': 'es_LA',
    'EE': 'et_EE', 'EG': 'ar_AR', 'ES': 'es_ES', 'ET': 'en_US', 'FI': 'fi_FI',
    'FJ': 'en_US', 'FR': 'fr_FR', 'GA': 'fr_FR', 'GB': 'en_GB', 'GE': 'ka_GE',
    'GH': 'en_US', 'GM': 'en_US', 'GN': 'fr_FR', 'GQ': 'es_LA', 'GR': 'el_GR',
    'GT': 'es_LA', 'GW': 'pt_PT', 'GY': 'en_US', 'HK': 'zh_HK', 'HN': 'es_LA',
    'HR': 'hr_HR', 'HT': 'fr_FR', 'HU': 'hu_HU', 'ID': 'id_ID', 'IE': 'en_GB',
    'IL': 'he_IL', 'IN': 'hi_IN', 'IQ': 'ar_AR', 'IR': 'fa_IR', 'IS': 'is_IS',
    'IT': 'it_IT', 'JM': 'en_US', 'JO': 'ar_AR', 'JP': 'ja_JP', 'KE': 'en_US',
    'KG': 'ky_KG', 'KH': 'km_KH', 'KM': 'fr_FR', 'KP': 'ko_KR', 'KR': 'ko_KR',
    'KW': 'ar_AR', 'KZ': 'kk_KZ', 'LA': 'en_US', 'LB': 'ar_AR', 'LI': 'de_DE',
    'LK': 'si_LK', 'LR': 'en_US', 'LS': 'en_US', 'LT': 'lt_LT', 'LU': 'fr_FR',
    'LV': 'lv_LV', 'LY': 'ar_AR', 'MA': 'ar_AR', 'MC': 'fr_FR', 'MD': 'ro_RO',
    'ME': 'sr_RS', 'MG': 'fr_FR', 'MK': 'mk_MK', 'ML': 'fr_FR', 'MM': 'my_MM',
    'MN': 'mn_MN', 'MR': 'ar_AR', 'MT': 'en_GB', 'MU': 'en_US', 'MV': 'en_US',
    'MW': 'en_US', 'MX': 'es_LA', 'MY': 'ms_MY', 'MZ': 'pt_PT', 'NA': 'en_US',
    'NE': 'fr_FR', 'NG': 'en_US', 'NI': 'es_LA', 'NL': 'nl_NL', 'NO': 'nb_NO',
    'NP': 'ne_NP', 'NZ': 'en_US', 'OM': 'ar_AR', 'PA': 'es_LA', 'PE': 'es_LA',
    'PH': 'en_US', 'PK': 'ur_PK', 'PL': 'pl_PL', 'PS': 'ar_AR', 'PT': 'pt_PT',
    'PY': 'es_LA', 'QA': 'ar_AR', 'RO': 'ro_RO', 'RS': 'sr_RS', 'RU': 'ru_RU',
    'RW': 'fr_FR', 'SA': 'ar_AR', 'SC': 'fr_FR', 'SD': 'ar_AR', 'SE': 'sv_SE',
    'SG': 'en_US', 'SI': 'sl_SI', 'SK': 'sk_SK', 'SL': 'en_US', 'SM': 'it_IT',
    'SN': 'fr_FR', 'SO': 'so_SO', 'SR': 'nl_NL', 'SS': 'en_US', 'ST': 'pt_PT',
    'SV': 'es_LA', 'SY': 'ar_AR', 'SZ': 'en_US', 'TD': 'fr_FR', 'TG': 'fr_FR',
    'TH': 'th_TH', 'TJ': 'tg_TJ', 'TL': 'pt_PT', 'TM': 'tk_TM', 'TN': 'ar_AR',
    'TO': 'en_US', 'TR': 'tr_TR', 'TT': 'en_US', 'TW': 'zh_TW', 'TZ': 'en_US',
    'UA': 'uk_UA', 'UG': 'en_US', 'US': 'en_US', 'UY': 'es_LA', 'UZ': 'uz_UZ',
    'VA': 'it_IT', 'VE': 'es_LA', 'VN': 'vi_VN', 'YE': 'ar_AR', 'ZA': 'en_US',
    'ZM': 'en_US', 'ZW': 'en_US',
}


def get_locale_code(cc):
    return COUNTRY_TO_LOCALE.get((cc or 'US').upper(), 'en_US')


def clamp_threads(value, has_proxies=False):
    try:
        value = int(value)
    except Exception:
        value = DEFAULT_THREADS
    max_limit = MAX_THREAD_LIMIT_WITH_PROXY if has_proxies else MAX_THREAD_LIMIT_NO_PROXY
    return max(1, min(value, max_limit))


def normalize_numbers(raw_numbers):
    seen = set()
    cleaned = []
    for raw in raw_numbers:
        s = re.sub(r'[^0-9]', '', str(raw).strip())
        if 7 <= len(s) <= 15 and s not in seen:
            seen.add(s)
            cleaned.append(s)
    return cleaned


def make_run_id():
    return time.strftime('%Y%m%d_%H%M%S')


def build_result_path(kind, cid, run_id):
    return os.path.join(RESULTS_DIR, f'{kind}_{cid}_{run_id}.txt')


# ============================================================
# PROXY & IP HELPERS
# ============================================================
free_proxies_cache = []
proxy_cache_lock = threading.Lock()


def fetch_new_ip():
    global free_proxies_cache
    with proxy_cache_lock:
        if not free_proxies_cache:
            try:
                r = requests.get(
                    'https://api.proxyscrape.com/v2/?request=displayproxies&protocol=http&timeout=3000&country=all&ssl=yes&anonymity=all',
                    timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
                    verify=False,
                )
                free_proxies_cache = [p.strip() for p in r.text.strip().split('\n') if p.strip()]
                random.shuffle(free_proxies_cache)
            except Exception:
                pass
        if free_proxies_cache:
            px = free_proxies_cache.pop(0)
            return {'http': f'http://{px}', 'https': f'http://{px}'}
    return None


def parse_proxy(proxy_str):
    if not proxy_str:
        return None
    if '://' not in proxy_str:
        if '@' in proxy_str:
            part1, part2 = proxy_str.split('@', 1)
            if '.' in part2.split(':')[0]:
                proxy_url = f'http://{part1}@{part2}'
            else:
                proxy_url = f'http://{part2}@{part1}'
        else:
            parts = proxy_str.split(':')
            if len(parts) == 4:
                if '.' in parts[0]:
                    ip, port, user, pwd = parts
                    proxy_url = f'http://{user}:{pwd}@{ip}:{port}'
                else:
                    user, pwd, ip, port = parts
                    proxy_url = f'http://{user}:{pwd}@{ip}:{port}'
            elif len(parts) == 2:
                proxy_url = f'http://{parts[0]}:{parts[1]}'
            else:
                return None
    else:
        proxy_url = proxy_str
    return {'http': proxy_url, 'https': proxy_url}


def test_proxy(proxies, domain=DEFAULT_SERVER):
    try:
        r = requests.get(
            f'https://{domain}',
            proxies=proxies,
            timeout=(5, 8),
            verify=False,
        )
        return r.status_code == 200
    except Exception:
        return False


def validate_proxies_parallel(proxy_lines, domain):
    parsed_list = []
    seen = set()
    for line in proxy_lines:
        parsed = parse_proxy(line)
        if parsed:
            key = parsed.get('http')
            if key not in seen:
                seen.add(key)
                parsed_list.append(parsed)

    if not parsed_list:
        return []

    workers = min(MAX_PROXY_TEST_WORKERS, len(parsed_list))
    working = []

    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {executor.submit(test_proxy, px, domain): px for px in parsed_list}
        for future in as_completed(future_map):
            px = future_map[future]
            try:
                ok = future.result()
            except Exception:
                ok = False
            if ok:
                working.append(px)
    return working


def get_ip_info(proxies=None):
    try:
        r = requests.get(
            'http://ip-api.com/json/',
            proxies=proxies,
            timeout=(4, 6),
            verify=False,
        )
        if r.status_code == 200:
            d = r.json()
            return {
                'country': d.get('country', 'Unknown'),
                'countryCode': d.get('countryCode', 'US'),
                'timezone': d.get('timezone', 'Unknown'),
            }
    except Exception:
        pass
    return {'country': 'Unknown', 'countryCode': 'US', 'timezone': 'Unknown'}


# ============================================================
# REQUEST HELPERS
# ============================================================
def browser_signature(browser_type):
    browser_type = browser_type or DEFAULT_BROWSER
    browser_map = {
        'Chrome': ('Chrome', '537.36'),
        'Firefox': ('Firefox', None),
        'Opera': ('OPR', '537.36'),
        'Edge': ('Edg', '537.36'),
        'Brave': ('Chrome', '537.36'),
        'Samsung': ('SamsungBrowser', '537.36'),
        'UC': ('UCBrowser', '537.36'),
        'DuckDuckGo': ('DuckDuckGo', '537.36'),
        'Vivaldi': ('Vivaldi', '537.36'),
        'Yandex': ('YaBrowser', '537.36'),
        'Kiwi': ('Kiwi', '537.36'),
        'Dolphin': ('Dolphin', '537.36'),
        'Mi Browser': ('MiuiBrowser', '537.36'),
        'Maxthon': ('MxBrowser', '537.36'),
        'Puffin': ('Puffin', '537.36'),
    }
    return browser_map.get(browser_type, ('Chrome', '537.36'))


def build_headers(locale='en_US', device_type='Random', browser_type='Chrome'):
    if device_type == 'Random':
        device_type = random.choice(['Android', 'iPhone'])

    browser_name, webkit = browser_signature(browser_type)

    if device_type == 'Android':
        andro_ver = random.choice(['10', '11', '12', '13', '14'])
        model = random.choice(['SM-G998B', 'SM-S908B', 'Pixel 6', 'Pixel 7', 'M2101K6G'])
        major_ver = random.randint(110, 126)
        if browser_name == 'Firefox':
            ua = f'Mozilla/5.0 (Android {andro_ver}; Mobile; rv:{major_ver}.0) Gecko/{major_ver}.0 Firefox/{major_ver}.0'
        else:
            webkit = webkit or '537.36'
            ua = f'Mozilla/5.0 (Linux; Android {andro_ver}; {model}) AppleWebKit/{webkit} (KHTML, like Gecko) {browser_name}/{major_ver}.0 Mobile Safari/{webkit}'
    elif device_type == 'iPhone':
        ios_ver = random.choice(['15_6_1', '16_0', '16_2', '16_5', '17_0', '17_1'])
        ios_main = ios_ver.split('_')[0]
        if browser_name == 'Firefox':
            ua = f'Mozilla/5.0 (iPhone; CPU iPhone OS {ios_ver} like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) FxiOS/{ios_main}.0 Mobile/15E148 Safari/605.1.15'
        else:
            ua = f'Mozilla/5.0 (iPhone; CPU iPhone OS {ios_ver} like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/{ios_main}.0 Mobile/15E148 Safari/604.1'
    elif device_type == 'KaiOS':
        ua = f'Mozilla/5.0 (Mobile; Nokia; rv:48.0) Gecko/48.0 Firefox/48.0 KaiOS/{random.choice(["2.5", "3.0", "3.1"])}'
    elif device_type == 'Windows Phone':
        ua = 'Mozilla/5.0 (Windows Phone 10.0; Android 7.0; Microsoft; Lumia 950) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Mobile Safari/537.36 Edge/15.15063'
    elif device_type == 'BlackBerry':
        ua = 'Mozilla/5.0 (BB10; Touch) AppleWebKit/537.10+ (KHTML, like Gecko) Version/10.0.0.1337 Mobile Safari/537.10+'
    else:
        model = random.choice(['SM-G998B', 'Pixel 6', 'M2101K6G'])
        major_ver = random.randint(110, 126)
        ua = f'Mozilla/5.0 (Linux; Android 12; {model}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{major_ver}.0.0.0 Mobile Safari/537.36'

    return {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'accept-language': f'{locale},en;q=0.9',
        'user-agent': ua,
        'upgrade-insecure-requests': '1',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'cache-control': 'max-age=0',
    }


def process_sms(resp_text):
    return 'id="contact_point_selector_form"' in resp_text and 'name="recover_method"' in resp_text


def create_http_session(proxy=None):
    session = requests.Session()
    session.verify = False
    session.headers.update({'Connection': 'keep-alive'})
    adapter = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=0)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    if proxy:
        session.proxies.update(proxy)
    return session


# ============================================================
# CORE CHECKER
# ============================================================
def check(number, proxy=None, locale='en_US', browser_type='Chrome', retry_count=0,
          server_domain=DEFAULT_SERVER, device_type='Random'):
    if server_domain == 'Random':
        server_domain = random.choice([v for v in SERVER_MAP.values() if v != 'Random'])

    session = create_http_session(proxy=proxy)
    base_headers = build_headers(locale=locale, device_type=device_type, browser_type=browser_type)

    try:
        git_fb = session.get(
            f'https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr',
            headers=base_headers,
            timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
        )

        try:
            lsd = re.search('name="lsd" value="(.*?)"', git_fb.text).group(1)
        except Exception:
            try:
                lsd = re.search(r'\["LSD",\[\],\{"token":"(.*?)"\}', git_fb.text).group(1)
            except Exception:
                lsd = ''

        try:
            jazoest = re.search('name="jazoest" value="(.*?)"', git_fb.text).group(1)
        except Exception:
            try:
                jazoest = re.search('"initSprinkleValue":"(.*?)"', git_fb.text).group(1)
            except Exception:
                jazoest = ''

        if not lsd or not jazoest:
            if retry_count < MAX_RETRIES:
                new_proxy = fetch_new_ip() if proxy else None
                time.sleep(random.uniform(MIN_JITTER, MAX_JITTER))
                return check(number, new_proxy or proxy, locale, browser_type, retry_count + 1, server_domain, device_type)
            return 'error', 'IP Rate Limited'

        payload = {'lsd': lsd, 'jazoest': jazoest, 'email': number, 'did_submit': 'Search'}
        post_headers = base_headers.copy()
        post_headers.update({
            'content-type': 'application/x-www-form-urlencoded',
            'origin': f'https://{server_domain}',
            'referer': f'https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0',
            'sec-fetch-site': 'same-origin',
        })
        url = f'https://{server_domain}/login/identify/?ctx=recover&c=%2Flogin%2F&search_attempts=1&ars=facebook_login&alternate_search=0'
        resp = session.post(
            url,
            data=payload,
            headers=post_headers,
            allow_redirects=True,
            timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
        )

        if 'id="login_identify_search_error_msg"' in resp.text:
            err_match = re.search('id="login_identify_search_error_msg"[^>]*>(.*?)</div>', resp.text, re.IGNORECASE | re.DOTALL)
            err_text = err_match.group(1).lower() if err_match else ''
            if any(k in err_text for k in ['temporarily blocked', 'try again', 'too many', 'limit', 'spam', 'unusual', 'restrict']):
                if retry_count < MAX_RETRIES:
                    new_proxy = fetch_new_ip() if proxy else None
                    return check(number, new_proxy or proxy, locale, browser_type, retry_count + 1, server_domain, device_type)
                return 'error', 'Soft Ban'
            return 'not_found', 'Account Not Found'

        if 'action="/login/identify/?ctx=recover' in resp.text:
            return 'found', 'Multiple Accounts Found'

        if resp.url.startswith(f'https://{server_domain}/login/account_recovery/name_search/'):
            resp2 = session.get(resp.url, headers=base_headers, timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT))
            if 'action="/login/account_recovery/name_search/?flow=initiate_view' in resp2.text:
                resp3 = session.get(
                    f'https://{server_domain}/recover/initiate/?c=%2Flogin%2F&fl=initiate_view&ctx=msite_initiate_view',
                    headers=base_headers,
                    timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
                )
                if process_sms(resp3.text):
                    return 'found', 'Account Found (SMS Option)'
                return 'found', 'Account Found (Other Option)'

        if resp.url.startswith(f'https://{server_domain}/login/device-based/ar/login/?ldata='):
            resp2 = session.get(resp.url, headers=base_headers, timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT))
            if 'id="contact_point_selector_form"' in resp2.text:
                if process_sms(resp2.text):
                    return 'found', 'Account Found (SMS Option)'
                return 'found', 'Account Found (Contact Point)'
            if 'name="captcha_response"' in resp2.text:
                return 'not_found', 'Captcha Block'
            if '/help/121104481304395' in resp2.text or '/help/103873106370583' in resp2.text:
                return 'found', 'Account Disabled (Found)'
            return 'error', 'Unknown Page'

        if 'window.MPageLoadClientMetrics' in resp.text:
            if retry_count < MAX_RETRIES:
                new_proxy = fetch_new_ip() if proxy else None
                return check(number, new_proxy or proxy, locale, browser_type, retry_count + 1, server_domain, device_type)
            return 'error', 'Bot Block'

        return 'error', 'Unknown Response'

    except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
        if retry_count < MAX_RETRIES:
            new_proxy = fetch_new_ip() if proxy else None
            return check(number, new_proxy or proxy, locale, browser_type, retry_count + 1, server_domain, device_type)
        return 'error', 'Network Error'
    except Exception as e:
        return 'error', str(e)[:80]
    finally:
        session.close()


# ============================================================
# FILE PARSERS
# ============================================================
def extract_numbers_from_excel(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet = wb.active
        target_col = None
        max_matches = 0
        sample_limit = min(25, sheet.max_row or 0)

        for col_idx in range(1, sheet.max_column + 1):
            match_count = 0
            for row_idx in range(2, sample_limit + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val:
                    s = re.sub(r'[^0-9]', '', str(val).strip())
                    if 7 <= len(s) <= 15:
                        match_count += 1
            if match_count > max_matches:
                max_matches = match_count
                target_col = col_idx

        if not target_col:
            wb.close()
            return None, 'No phone column found'

        numbers = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=target_col, max_col=target_col, values_only=True):
            val = row[0]
            if val:
                s = re.sub(r'[^0-9]', '', str(val).strip())
                if 7 <= len(s) <= 15:
                    numbers.append(s)

        wb.close()
        return normalize_numbers(numbers), None
    except Exception as e:
        return None, str(e)


# ============================================================
# TELEGRAM SAFETY HELPERS
# ============================================================
def safe_send(chat_id, text, reply_markup=None, retries=4):
    for attempt in range(retries):
        try:
            return bot.send_message(chat_id, text, reply_markup=reply_markup)
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1.5 * (attempt + 1))
            else:
                print(f'safe_send failed after {retries} attempts: {e}')
    return None


def safe_edit(chat_id, msg_id, text, reply_markup=None, retries=2):
    for attempt in range(retries):
        try:
            bot.edit_message_text(text, chat_id, msg_id, reply_markup=reply_markup)
            return True
        except Exception as e:
            err = str(e).lower()
            if 'message is not modified' in err:
                return True
            if attempt < retries - 1:
                time.sleep(1)
    return False


def format_elapsed(seconds):
    seconds = max(0, int(seconds))
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    if h:
        return f'{h}h {m}m {s}s'
    if m:
        return f'{m}m {s}s'
    return f'{s}s'


# ============================================================
# RUNNER
# ============================================================
def run_checker(cid, chat_id):
    s = get_session(cid)
    numbers = list(s['numbers'])
    server = s['server']
    device = s['device']
    browser = s['browser']
    proxies = list(s['proxy_list'])
    threads = clamp_threads(s['threads'], has_proxies=bool(proxies))
    run_id = make_run_id()

    with session_lock:
        s['threads'] = threads
        s['run_id'] = run_id
        s['stats'] = {'checked': 0, 'found': 0, 'not_found': 0, 'error': 0}
        s['found_numbers'] = []
        s['not_found_numbers'] = []
        s['error_numbers'] = []
        s['result_details'] = {}
        s['last_progress_text'] = None
        s['last_progress_update'] = 0
        s['start_ts'] = time.time()
        s['active_workers'] = 0
        s['last_result_paths'] = {
            'found': build_result_path('found', cid, run_id),
            'not_found': build_result_path('not_found', cid, run_id),
            'error': build_result_path('error', cid, run_id),
        }

    locale = 'en_US'
    if proxies:
        try:
            ip_info = get_ip_info(proxies[0])
            locale = get_locale_code(ip_info.get('countryCode'))
        except Exception:
            locale = 'en_US'
    with session_lock:
        s['current_locale'] = locale

    total = len(numbers)

    def send_progress_update(force=False):
        with session_lock:
            st = dict(s['stats'])
            running = s['running']
            progress_msg_id = s['progress_msg_id']
            last_progress_text = s.get('last_progress_text')
            last_progress_update = s.get('last_progress_update', 0)
            active_workers = s.get('active_workers', 0)
            started = s.get('start_ts') or time.time()

        now = time.time()
        if not force and now - last_progress_update < PROGRESS_UPDATE_INTERVAL:
            return

        pct = int((st['checked'] / total) * 100) if total > 0 else 0
        filled = int(pct / 5)
        bar = '█' * filled + '░' * (20 - filled)
        elapsed = now - started
        speed = (st['checked'] / elapsed) if elapsed > 0 else 0
        eta = int((total - st['checked']) / speed) if speed > 0 and st['checked'] < total else 0

        text = (
            f"⚡ <b>Checking in Progress...</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 Progress : [{bar}] {pct}%\n"
            f"🔢 Checked  : <code>{st['checked']}/{total}</code>\n"
            f"✅ Found    : <code>{st['found']}</code>\n"
            f"❌ Not Found: <code>{st['not_found']}</code>\n"
            f"⚠️ Error    : <code>{st['error']}</code>\n"
            f"⚙️ Workers  : <code>{active_workers}/{threads}</code>\n"
            f"🚀 Speed    : <code>{speed:.2f}/sec</code>\n"
            f"⏱ Elapsed  : <code>{format_elapsed(elapsed)}</code>\n"
            f"🕒 ETA      : <code>{format_elapsed(eta) if eta else '0s'}</code>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🌐 Server   : <code>{server}</code>\n"
            f"📱 Device   : <code>{device}</code>\n"
            f"🌍 Browser  : <code>{browser}</code>\n"
            f"🔑 Proxies  : <code>{len(proxies)}</code>"
        )

        if not force and text == last_progress_text:
            with session_lock:
                s['last_progress_update'] = now
            return

        stop_mk = InlineKeyboardMarkup()
        if running:
            stop_mk.add(InlineKeyboardButton('🛑 Stop Checking', callback_data='stop_check'))

        ok = False
        if progress_msg_id:
            ok = safe_edit(chat_id, progress_msg_id, text, stop_mk)
        if not ok:
            msg = safe_send(chat_id, text, stop_mk)
            if msg:
                with session_lock:
                    s['progress_msg_id'] = msg.message_id

        with session_lock:
            s['last_progress_text'] = text
            s['last_progress_update'] = now

    init_msg = safe_send(
        chat_id,
        f"🚀 <b>Starting Checker...</b>\n"
        f"📋 Numbers: <code>{total}</code>\n"
        f"🌐 Server : <code>{server}</code>\n"
        f"📱 Device : <code>{device}</code>\n"
        f"🌍 Browser: <code>{browser}</code>\n"
        f"🔢 Threads: <code>{threads}</code>\n"
        f"🔄 Refresh: <code>{PROGRESS_UPDATE_INTERVAL}s</code>"
    )
    if init_msg:
        with session_lock:
            s['progress_msg_id'] = init_msg.message_id

    send_progress_update(force=True)

    def worker(number):
        with session_lock:
            if s['stop_flag']:
                return
            s['active_workers'] += 1

        proxy = random.choice(proxies) if proxies else None
        try:
            result, msg_txt = check(
                number,
                proxy=proxy,
                locale=locale,
                browser_type=browser,
                server_domain=server,
                device_type=device,
            )
        except Exception as e:
            result, msg_txt = 'error', str(e)[:80]

        with session_lock:
            s['active_workers'] -= 1
            s['stats']['checked'] += 1
            s['result_details'][number] = msg_txt
            if result == 'found':
                s['stats']['found'] += 1
                s['found_numbers'].append(number)
            elif result == 'not_found':
                s['stats']['not_found'] += 1
                s['not_found_numbers'].append(number)
            else:
                s['stats']['error'] += 1
                s['error_numbers'].append(number)

    futures = set()
    try:
        with ThreadPoolExecutor(max_workers=threads) as executor:
            num_iter = iter(numbers)

            while True:
                with session_lock:
                    should_stop = s['stop_flag']

                while not should_stop and len(futures) < threads:
                    try:
                        number = next(num_iter)
                    except StopIteration:
                        break
                    futures.add(executor.submit(worker, number))

                if not futures:
                    break

                done, futures = wait(futures, timeout=PROGRESS_UPDATE_INTERVAL, return_when=FIRST_COMPLETED)
                send_progress_update(force=True)

                if should_stop:
                    for future in futures:
                        future.cancel()
                    break
    finally:
        with session_lock:
            s['running'] = False
            s['stop_flag'] = False
            s['active_workers'] = 0

    try:
        found_path = s['last_result_paths']['found']
        not_found_path = s['last_result_paths']['not_found']
        error_path = s['last_result_paths']['error']

        with open(found_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(s['found_numbers']))
        with open(not_found_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(s['not_found_numbers']))
        with open(error_path, 'w', encoding='utf-8') as f:
            lines = [f"{num} | {s['result_details'].get(num, 'Unknown Error')}" for num in s['error_numbers']]
            f.write('\n'.join(lines))
    except Exception:
        pass

    st = s['stats']
    pct = int((st['checked'] / total) * 100) if total > 0 else 0
    elapsed = time.time() - (s.get('start_ts') or time.time())
    speed = (st['checked'] / elapsed) if elapsed > 0 else 0
    summary = (
        f"✅ <b>Checking Complete!</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🔢 Total    : <code>{total}</code>\n"
        f"📊 Checked  : <code>{st['checked']}</code> ({pct}%)\n"
        f"✅ Found    : <code>{st['found']}</code>\n"
        f"❌ Not Found: <code>{st['not_found']}</code>\n"
        f"⚠️ Error    : <code>{st['error']}</code>\n"
        f"🚀 Avg Speed: <code>{speed:.2f}/sec</code>\n"
        f"⏱ Time     : <code>{format_elapsed(elapsed)}</code>\n"
        f"🆔 Run ID   : <code>{run_id}</code>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )

    result_mk = InlineKeyboardMarkup(row_width=2)
    result_mk.add(
        InlineKeyboardButton('📥 Download Found', callback_data='dl_found'),
        InlineKeyboardButton('📥 Download Not Found', callback_data='dl_not_found'),
    )
    if s['error_numbers']:
        result_mk.add(InlineKeyboardButton('📥 Download Errors', callback_data='dl_error'))
    result_mk.add(InlineKeyboardButton('🏠 Main Menu', callback_data='main_menu'))

    if s['progress_msg_id']:
        ok = safe_edit(chat_id, s['progress_msg_id'], summary, result_mk)
        if not ok:
            safe_send(chat_id, summary, result_mk)
    else:
        safe_send(chat_id, summary, result_mk)


# ============================================================
# UI BUILDERS
# ============================================================
def main_menu_text(cid):
    s = get_session(cid)
    status = '🟢 Running' if s['running'] else '🔴 Idle'
    runtime_threads = clamp_threads(s['threads'], has_proxies=bool(s['proxy_list']))
    return (
        f"🤖 <b>Checker Bot</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📌 Status  : {status}\n"
        f"🌐 Server  : <code>{s['server']}</code>\n"
        f"📱 Device  : <code>{s['device']}</code>\n"
        f"🌍 Browser : <code>{s['browser']}</code>\n"
        f"🔢 Threads : <code>{s['threads']}</code>\n"
        f"⚡ Runtime  : <code>{runtime_threads}</code>\n"
        f"🔑 Proxies : <code>{len(s['proxy_list'])}</code>\n"
        f"🔄 Refresh : <code>{PROGRESS_UPDATE_INTERVAL}s</code>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    )


def main_menu_markup(cid):
    s = get_session(cid)
    mk = InlineKeyboardMarkup(row_width=2)
    mk.add(
        InlineKeyboardButton('🚀 Start Check', callback_data='start_check'),
        InlineKeyboardButton('⚙️ Settings', callback_data='settings'),
    )
    mk.add(
        InlineKeyboardButton('📊 Live Stats', callback_data='live_stats'),
        InlineKeyboardButton('📁 Download Results', callback_data='download_results'),
    )
    if s['running']:
        mk.add(InlineKeyboardButton('🛑 Stop Checking', callback_data='stop_check'))
    return mk


def settings_markup():
    mk = InlineKeyboardMarkup(row_width=2)
    mk.add(
        InlineKeyboardButton('🌐 Server', callback_data='set_server'),
        InlineKeyboardButton('📱 Device', callback_data='set_device'),
    )
    mk.add(
        InlineKeyboardButton('🌍 Browser', callback_data='set_browser'),
        InlineKeyboardButton('🔢 Threads', callback_data='set_threads'),
    )
    mk.add(
        InlineKeyboardButton('🔑 Add Proxy', callback_data='set_proxy'),
        InlineKeyboardButton('🗑 Clear Proxy', callback_data='clear_proxy'),
    )
    mk.add(InlineKeyboardButton('🔄 Reset All Config', callback_data='reset_config'))
    mk.add(InlineKeyboardButton('🏠 Back to Menu', callback_data='main_menu'))
    return mk


def server_markup():
    mk = InlineKeyboardMarkup(row_width=2)
    for k, v in SERVER_MAP.items():
        label = '🎲 Random (Mix)' if v == 'Random' else v
        mk.add(InlineKeyboardButton(label, callback_data=f'sv_{k}'))
    mk.add(InlineKeyboardButton('🔙 Back', callback_data='settings'))
    return mk


def device_markup():
    mk = InlineKeyboardMarkup(row_width=2)
    labels = {'0': '🎲 Random', '1': '🤖 Android', '2': '🍎 iPhone', '3': '📱 KaiOS', '4': '🪟 Windows Phone', '5': '🫐 BlackBerry'}
    for k, label in labels.items():
        mk.add(InlineKeyboardButton(label, callback_data=f'dv_{k}'))
    mk.add(InlineKeyboardButton('🔙 Back', callback_data='settings'))
    return mk


def browser_markup():
    mk = InlineKeyboardMarkup(row_width=3)
    icons = {
        '0': '🎲 Random', '1': '🟡 Chrome', '2': '🦊 Firefox', '3': '🔴 Opera',
        '4': '💙 Edge', '5': '🦁 Brave', '6': '📱 Samsung', '7': '🟠 UC',
        '8': '🦆 DuckDuckGo', '9': '🟣 Vivaldi', '10': '🅨 Yandex',
        '11': '🥝 Kiwi', '12': '🐬 Dolphin', '13': '🔵 Mi Browser',
        '14': '⚡ Maxthon', '15': '🌊 Puffin'
    }
    btns = [InlineKeyboardButton(label, callback_data=f'br_{k}') for k, label in icons.items()]
    mk.add(*btns)
    mk.add(InlineKeyboardButton('🔙 Back', callback_data='settings'))
    return mk


# ============================================================
# HANDLERS
# ============================================================
@bot.message_handler(commands=['myid'])
def cmd_myid(message):
    uid = message.from_user.id
    bot.send_message(
        message.chat.id,
        f"🆔 <b>Your Telegram User ID:</b>\n<code>{uid}</code>\n\n"
        f"👉 checker_bot.py তে এই লাইনে এই ID বসাও:\n"
        f"<code>ADMIN_IDS = set([{uid}])</code>"
    )


@bot.message_handler(commands=['start'])
def cmd_start(message):
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, '❌ Access Denied.')
        return
    cid = message.chat.id
    get_session(cid)
    bot.send_message(cid, main_menu_text(cid), reply_markup=main_menu_markup(cid))


@bot.message_handler(content_types=['document'])
def handle_file(message):
    if not is_admin(message.from_user.id):
        return
    cid = message.chat.id
    s = get_session(cid)

    if s.get('awaiting') != 'file':
        return

    s['awaiting'] = None
    doc = message.document
    fname = (doc.file_name or '').lower()

    try:
        file_info = bot.get_file(doc.file_id)
        downloaded = bot.download_file(file_info.file_path)
    except Exception as e:
        bot.send_message(cid, f'❌ File download failed: {e}')
        return

    numbers = []
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        tmp_path = f'/tmp/numlist_{cid}.xlsx'
        with open(tmp_path, 'wb') as f:
            f.write(downloaded)
        numbers, err = extract_numbers_from_excel(tmp_path)
        if err:
            bot.send_message(cid, f'❌ Excel parse error: {err}')
            return
    elif fname.endswith('.txt'):
        try:
            content = downloaded.decode('utf-8', errors='ignore')
            numbers = normalize_numbers(content.splitlines())
        except Exception as e:
            bot.send_message(cid, f'❌ Text file parse error: {e}')
            return
    else:
        bot.send_message(cid, '❌ Only <b>.txt</b> or <b>.xlsx</b> files are supported.')
        return

    if not numbers:
        bot.send_message(cid, '⚠️ No valid numbers found in the file.')
        return

    s['numbers'] = numbers
    bot.send_message(
        cid,
        f"✅ <b>File loaded!</b>\n"
        f"📋 Unique valid numbers: <code>{len(numbers)}</code>\n\n"
        f"Ready to start. Use the button below.",
        reply_markup=InlineKeyboardMarkup().add(
            InlineKeyboardButton('🚀 Start Now!', callback_data='begin_check')
        )
    )


@bot.message_handler(func=lambda m: True)
def handle_text(message):
    if not is_admin(message.from_user.id):
        return

    cid = message.chat.id
    s = get_session(cid)
    text = (message.text or '').strip()

    if s.get('awaiting') == 'threads':
        s['awaiting'] = None
        try:
            t = int(text)
            if not (1 <= t <= 2000):
                raise ValueError
            s['threads'] = t
            actual = clamp_threads(t, has_proxies=bool(s['proxy_list']))
            note = '' if actual == t else f"\n⚠️ Runtime cap applied: <code>{actual}</code> for stability"
            bot.send_message(cid, f"✅ Threads set to <code>{t}</code>{note}", reply_markup=settings_markup())
        except Exception:
            bot.send_message(cid, '❌ Invalid! Enter a number between 1–2000.', reply_markup=settings_markup())

    elif s.get('awaiting') == 'proxy':
        s['awaiting'] = None
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        wait_msg = bot.send_message(cid, f'⏳ Testing {len(lines)} proxy/proxies...')

        working = validate_proxies_parallel(lines, s['server'])

        existing = {px.get('http') for px in s['proxy_list']}
        added = 0
        for px in working:
            key = px.get('http')
            if key not in existing:
                s['proxy_list'].append(px)
                existing.add(key)
                added += 1

        try:
            bot.edit_message_text(
                f"✅ <b>{added}/{len(lines)}</b> proxies added and working!\n"
                f"🔑 Total proxies: <code>{len(s['proxy_list'])}</code>",
                cid,
                wait_msg.message_id,
                reply_markup=settings_markup()
            )
        except Exception:
            bot.send_message(cid, f'✅ {added} proxies added.', reply_markup=settings_markup())

    elif s.get('awaiting') == 'numbers_text':
        s['awaiting'] = None
        nums = normalize_numbers(text.splitlines())
        if nums:
            s['numbers'] = nums
            bot.send_message(
                cid,
                f"✅ <b>{len(nums)} unique valid numbers loaded!</b>",
                reply_markup=InlineKeyboardMarkup().add(
                    InlineKeyboardButton('🚀 Start Now!', callback_data='begin_check')
                )
            )
        else:
            bot.send_message(cid, '⚠️ No valid numbers found.')


@bot.callback_query_handler(func=lambda call: True)
def handle_callback(call):
    if not is_admin(call.from_user.id):
        return

    cid = call.message.chat.id
    mid = call.message.message_id
    data = call.data
    s = get_session(cid)

    bot.answer_callback_query(call.id)

    if data == 'main_menu':
        try:
            bot.edit_message_text(main_menu_text(cid), cid, mid, reply_markup=main_menu_markup(cid))
        except Exception:
            bot.send_message(cid, main_menu_text(cid), reply_markup=main_menu_markup(cid))

    elif data == 'settings':
        actual_threads = clamp_threads(s['threads'], has_proxies=bool(s['proxy_list']))
        s_txt = (
            f"⚙️ <b>Settings</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🌐 Server  : <code>{s['server']}</code>\n"
            f"📱 Device  : <code>{s['device']}</code>\n"
            f"🌍 Browser : <code>{s['browser']}</code>\n"
            f"🔢 Threads : <code>{s['threads']}</code>\n"
            f"⚡ Runtime  : <code>{actual_threads}</code>\n"
            f"🔑 Proxies : <code>{len(s['proxy_list'])}</code>"
        )
        try:
            bot.edit_message_text(s_txt, cid, mid, reply_markup=settings_markup())
        except Exception:
            bot.send_message(cid, s_txt, reply_markup=settings_markup())

    elif data == 'set_server':
        try:
            bot.edit_message_text('🌐 <b>Select Server:</b>', cid, mid, reply_markup=server_markup())
        except Exception:
            bot.send_message(cid, '🌐 <b>Select Server:</b>', reply_markup=server_markup())

    elif data.startswith('sv_'):
        key = data[3:]
        s['server'] = SERVER_MAP.get(key, DEFAULT_SERVER)
        bot.answer_callback_query(call.id, f"✅ Server: {s['server']}", show_alert=True)
        try:
            bot.edit_message_text(
                f"⚙️ <b>Settings</b>\n🌐 Server: <code>{s['server']}</code>",
                cid, mid, reply_markup=settings_markup()
            )
        except Exception:
            bot.send_message(cid, f"✅ Server set to <code>{s['server']}</code>", reply_markup=settings_markup())

    elif data == 'set_device':
        try:
            bot.edit_message_text('📱 <b>Select Device:</b>', cid, mid, reply_markup=device_markup())
        except Exception:
            bot.send_message(cid, '📱 <b>Select Device:</b>', reply_markup=device_markup())

    elif data.startswith('dv_'):
        key = data[3:]
        s['device'] = DEVICE_MAP.get(key, DEFAULT_DEVICE)
        bot.answer_callback_query(call.id, f"✅ Device: {s['device']}", show_alert=True)
        try:
            bot.edit_message_text(
                f"⚙️ <b>Settings</b>\n📱 Device: <code>{s['device']}</code>",
                cid, mid, reply_markup=settings_markup()
            )
        except Exception:
            bot.send_message(cid, f"✅ Device set to <code>{s['device']}</code>", reply_markup=settings_markup())

    elif data == 'set_browser':
        try:
            bot.edit_message_text('🌍 <b>Select Browser:</b>', cid, mid, reply_markup=browser_markup())
        except Exception:
            bot.send_message(cid, '🌍 <b>Select Browser:</b>', reply_markup=browser_markup())

    elif data.startswith('br_'):
        key = data[3:]
        s['browser'] = BROWSER_MAP.get(key, DEFAULT_BROWSER)
        bot.answer_callback_query(call.id, f"✅ Browser: {s['browser']}", show_alert=True)
        try:
            bot.edit_message_text(
                f"⚙️ <b>Settings</b>\n🌍 Browser: <code>{s['browser']}</code>",
                cid, mid, reply_markup=settings_markup()
            )
        except Exception:
            bot.send_message(cid, f"✅ Browser set to <code>{s['browser']}</code>", reply_markup=settings_markup())

    elif data == 'set_threads':
        mk_t = InlineKeyboardMarkup(row_width=3)
        mk_t.add(
            InlineKeyboardButton('⚡ 25', callback_data='th_25'),
            InlineKeyboardButton('⚡ 50', callback_data='th_50'),
            InlineKeyboardButton('🚀 100', callback_data='th_100'),
            InlineKeyboardButton('💥 150', callback_data='th_150'),
            InlineKeyboardButton('🔥 200', callback_data='th_200'),
            InlineKeyboardButton('🌪 300', callback_data='th_300'),
        )
        mk_t.add(InlineKeyboardButton('✏️ Custom (type)', callback_data='th_custom'))
        mk_t.add(InlineKeyboardButton('🔙 Back', callback_data='settings'))
        txt = (
            f"🔢 <b>Set Thread Count</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"Current : <code>{s['threads']}</code>\n"
            f"Runtime : <code>{clamp_threads(s['threads'], has_proxies=bool(s['proxy_list']))}</code>\n"
            f"Range   : 1 – 2000\n\n"
            f"📌 <b>Recommended:</b>\n"
            f"• No Proxy      → 25–100\n"
            f"• With Proxy    → 100–300\n"
            f"• Refresh Rate  → every {PROGRESS_UPDATE_INTERVAL}s"
        )
        try:
            bot.edit_message_text(txt, cid, mid, reply_markup=mk_t)
        except Exception:
            bot.send_message(cid, txt, reply_markup=mk_t)

    elif data.startswith('th_'):
        key = data[3:]
        if key == 'custom':
            s['awaiting'] = 'threads'
            try:
                bot.edit_message_text(
                    f"✏️ <b>Custom Thread Count</b>\n\nCurrent: <code>{s['threads']}</code>\nRange: 1–2000\n\nSend the number:",
                    cid, mid
                )
            except Exception:
                bot.send_message(cid, '✏️ Send thread count (1–2000):')
        else:
            t = int(key)
            s['threads'] = t
            actual = clamp_threads(t, has_proxies=bool(s['proxy_list']))
            msg = f'✅ Threads set to {t}'
            if actual != t:
                msg += f'\nRuntime cap: {actual}'
            bot.answer_callback_query(call.id, msg, show_alert=True)
            try:
                bot.edit_message_text(
                    f"⚙️ <b>Settings</b>\n🔢 Threads: <code>{t}</code>\n⚡ Runtime: <code>{actual}</code>",
                    cid, mid, reply_markup=settings_markup()
                )
            except Exception:
                bot.send_message(cid, f"✅ Threads set to <code>{t}</code>\n⚡ Runtime: <code>{actual}</code>", reply_markup=settings_markup())

    elif data == 'set_proxy':
        s['awaiting'] = 'proxy'
        try:
            bot.edit_message_text(
                '🔑 <b>Add Proxy</b>\n\nSend one or multiple proxies, one per line.\n'
                'Supported formats:\n'
                '<code>ip:port</code>\n'
                '<code>user:pass@ip:port</code>\n'
                '<code>http://user:pass@ip:port</code>',
                cid, mid
            )
        except Exception:
            bot.send_message(
                cid,
                '✏️ Send proxy/proxies (one per line):\n'
                '<code>ip:port</code> or <code>user:pass@ip:port</code>'
            )

    elif data == 'clear_proxy':
        s['proxy_list'] = []
        bot.answer_callback_query(call.id, '✅ All proxies cleared!', show_alert=True)
        try:
            bot.edit_message_text(
                '⚙️ <b>Settings</b>\n🔑 Proxies: <code>0</code> (cleared)',
                cid, mid, reply_markup=settings_markup()
            )
        except Exception:
            bot.send_message(cid, '✅ Proxies cleared.', reply_markup=settings_markup())

    elif data == 'reset_config':
        mk_confirm = InlineKeyboardMarkup(row_width=2)
        mk_confirm.add(
            InlineKeyboardButton('✅ Yes, Reset', callback_data='reset_confirm'),
            InlineKeyboardButton('❌ Cancel', callback_data='settings'),
        )
        try:
            bot.edit_message_text(
                '⚠️ <b>Reset All Settings?</b>\n\n'
                f'• Server → {DEFAULT_SERVER}\n'
                f'• Device → {DEFAULT_DEVICE}\n'
                f'• Browser → {DEFAULT_BROWSER}\n'
                f'• Threads → {DEFAULT_THREADS}\n'
                '• Proxy → Clear all\n\n'
                'Numbers list will NOT be cleared.',
                cid, mid, reply_markup=mk_confirm
            )
        except Exception:
            bot.send_message(cid, '⚠️ Confirm reset?', reply_markup=mk_confirm)

    elif data == 'reset_confirm':
        if s['running']:
            bot.answer_callback_query(call.id, '⚠️ Stop checking first!', show_alert=True)
            return
        numbers_keep = s.get('numbers', [])
        sessions[cid] = create_default_session()
        sessions[cid]['numbers'] = numbers_keep
        bot.answer_callback_query(call.id, '✅ All settings reset to default!', show_alert=True)
        s2 = get_session(cid)
        s_txt = (
            f"🔄 <b>Config Reset Done!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🌐 Server  : <code>{s2['server']}</code>\n"
            f"📱 Device  : <code>{s2['device']}</code>\n"
            f"🌍 Browser : <code>{s2['browser']}</code>\n"
            f"🔢 Threads : <code>{s2['threads']}</code>\n"
            f"🔑 Proxies : <code>0</code>"
        )
        try:
            bot.edit_message_text(s_txt, cid, mid, reply_markup=settings_markup())
        except Exception:
            bot.send_message(cid, s_txt, reply_markup=settings_markup())

    elif data == 'start_check':
        if s['running']:
            bot.answer_callback_query(call.id, '⚠️ Already running! Stop first.', show_alert=True)
            return
        mk = InlineKeyboardMarkup(row_width=1)
        mk.add(
            InlineKeyboardButton('📁 Upload .txt or .xlsx File', callback_data='upload_file'),
            InlineKeyboardButton('✏️ Paste Numbers (Text)', callback_data='paste_numbers'),
        )
        if s.get('numbers'):
            mk.add(InlineKeyboardButton(f"▶️ Use Last List ({len(s['numbers'])} numbers)", callback_data='begin_check'))
        mk.add(InlineKeyboardButton('🔙 Back', callback_data='main_menu'))
        try:
            bot.edit_message_text('📂 <b>How to load numbers?</b>', cid, mid, reply_markup=mk)
        except Exception:
            bot.send_message(cid, '📂 <b>How to load numbers?</b>', reply_markup=mk)

    elif data == 'upload_file':
        s['awaiting'] = 'file'
        try:
            bot.edit_message_text(
                '📤 <b>Send your file</b>\n\nSupported: <b>.txt</b> (one number per line) or <b>.xlsx</b>',
                cid, mid
            )
        except Exception:
            bot.send_message(cid, '📤 Send your .txt or .xlsx file now:')

    elif data == 'paste_numbers':
        s['awaiting'] = 'numbers_text'
        try:
            bot.edit_message_text(
                '✏️ <b>Paste numbers</b>\n\nSend numbers, one per line:\n<code>9920000001\n9920000002\n...</code>',
                cid, mid
            )
        except Exception:
            bot.send_message(cid, '✏️ Send numbers, one per line:')

    elif data == 'begin_check':
        if s['running']:
            bot.answer_callback_query(call.id, '⚠️ Already running!', show_alert=True)
            return
        if not s.get('numbers'):
            bot.answer_callback_query(call.id, '❌ No numbers loaded! Upload a file first.', show_alert=True)
            return
        s['numbers'] = normalize_numbers(s['numbers'])
        s['running'] = True
        s['stop_flag'] = False
        s['progress_msg_id'] = None
        t = threading.Thread(target=run_checker, args=(cid, cid), daemon=True)
        t.start()

    elif data == 'stop_check':
        if s['running']:
            s['stop_flag'] = True
            bot.answer_callback_query(call.id, '🛑 Stop signal sent...', show_alert=True)
            try:
                bot.edit_message_text(
                    '⏹ <b>Stopping...</b>\nWaiting for current checks to finish.',
                    cid, mid
                )
            except Exception:
                bot.send_message(cid, '⏹ Stopping...')
        else:
            bot.answer_callback_query(call.id, 'ℹ️ Nothing is running.', show_alert=True)

    elif data == 'live_stats':
        st = s['stats']
        total = len(s['numbers']) if s['numbers'] else 0
        pct = int((st['checked'] / total) * 100) if total > 0 else 0
        status_txt = '🟢 Running' if s['running'] else '🔴 Idle'
        elapsed = time.time() - (s.get('start_ts') or time.time()) if s['running'] else 0
        speed = (st['checked'] / elapsed) if elapsed > 0 else 0
        stats_text = (
            f"📊 <b>Live Stats</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📌 Status    : {status_txt}\n"
            f"📋 Total     : <code>{total}</code>\n"
            f"🔢 Checked   : <code>{st['checked']}</code> ({pct}%)\n"
            f"✅ Found     : <code>{st['found']}</code>\n"
            f"❌ Not Found : <code>{st['not_found']}</code>\n"
            f"⚠️ Error     : <code>{st['error']}</code>\n"
            f"🚀 Speed     : <code>{speed:.2f}/sec</code>\n"
            f"⏱ Elapsed   : <code>{format_elapsed(elapsed)}</code>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━"
        )
        mk = InlineKeyboardMarkup(row_width=2)
        mk.add(
            InlineKeyboardButton('🔄 Refresh', callback_data='live_stats'),
            InlineKeyboardButton('🏠 Menu', callback_data='main_menu'),
        )
        try:
            bot.edit_message_text(stats_text, cid, mid, reply_markup=mk)
        except Exception:
            bot.send_message(cid, stats_text, reply_markup=mk)

    elif data == 'download_results':
        paths = s.get('last_result_paths', {})
        found_path = paths.get('found')
        nf_path = paths.get('not_found')
        err_path = paths.get('error')
        has_found = bool(found_path and os.path.exists(found_path) and os.path.getsize(found_path) > 0)
        has_nf = bool(nf_path and os.path.exists(nf_path) and os.path.getsize(nf_path) > 0)
        has_err = bool(err_path and os.path.exists(err_path) and os.path.getsize(err_path) > 0)
        if not has_found and not has_nf and not has_err:
            bot.answer_callback_query(call.id, 'ℹ️ No results yet. Run a check first.', show_alert=True)
            return
        mk = InlineKeyboardMarkup(row_width=2)
        if has_found:
            mk.add(InlineKeyboardButton(f"✅ Found ({len(s['found_numbers'])})", callback_data='dl_found'))
        if has_nf:
            mk.add(InlineKeyboardButton(f"❌ Not Found ({len(s['not_found_numbers'])})", callback_data='dl_not_found'))
        if has_err:
            mk.add(InlineKeyboardButton(f"⚠️ Errors ({len(s['error_numbers'])})", callback_data='dl_error'))
        mk.add(InlineKeyboardButton('🏠 Menu', callback_data='main_menu'))
        try:
            bot.edit_message_text('📁 <b>Download Results:</b>', cid, mid, reply_markup=mk)
        except Exception:
            bot.send_message(cid, '📁 <b>Download Results:</b>', reply_markup=mk)

    elif data == 'dl_found':
        path = s.get('last_result_paths', {}).get('found')
        if path and os.path.exists(path) and os.path.getsize(path) > 0:
            with open(path, 'rb') as f:
                bot.send_document(cid, f, caption=f"✅ Found: <code>{len(s['found_numbers'])}</code> numbers")
        else:
            bot.answer_callback_query(call.id, 'ℹ️ No found numbers yet.', show_alert=True)

    elif data == 'dl_not_found':
        path = s.get('last_result_paths', {}).get('not_found')
        if path and os.path.exists(path) and os.path.getsize(path) > 0:
            with open(path, 'rb') as f:
                bot.send_document(cid, f, caption=f"❌ Not Found: <code>{len(s['not_found_numbers'])}</code> numbers")
        else:
            bot.answer_callback_query(call.id, 'ℹ️ No not-found numbers yet.', show_alert=True)

    elif data == 'dl_error':
        path = s.get('last_result_paths', {}).get('error')
        if path and os.path.exists(path) and os.path.getsize(path) > 0:
            with open(path, 'rb') as f:
                bot.send_document(cid, f, caption=f"⚠️ Errors: <code>{len(s['error_numbers'])}</code> numbers")
        else:
            bot.answer_callback_query(call.id, 'ℹ️ No error results yet.', show_alert=True)


# ============================================================
# RUN
# ============================================================
if __name__ == '__main__':
    print('=' * 45)
    print('  ✅ Bot is running!')
    print('  📌 /start  — Main menu')
    print('  📌 /myid   — Get your Telegram user ID')
    print(f'  🔄 Progress refresh every {PROGRESS_UPDATE_INTERVAL}s')
    if not ADMIN_IDS:
        print('  ⚠️ ADMIN_IDS is empty — all users have access!')
        print('     Send /myid to bot, then set your ID.')
    else:
        print(f'  🔐 Admins: {ADMIN_IDS}')
    print('=' * 45)

    while True:
        try:
            bot.infinity_polling(
                timeout=60,
                long_polling_timeout=30,
                restart_on_change=False,
                skip_pending=True,
            )
        except Exception as e:
            print(f'⚠️ Polling error: {e} — restarting in 5s...')
            time.sleep(5)
